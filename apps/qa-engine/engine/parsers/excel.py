"""Excel (XLSX/XLS) ingestion with sheet/row/cell source metadata (FR-IN-008).

Each worksheet becomes segments preserving worksheet name, row number and
column references so generated requirements can cite workbook, worksheet and
cell/row source (FR-IN-008, §11 file-parsing integration).
"""

from __future__ import annotations

import io

from engine.contracts.schemas import ParsedSegment

MAX_ROWS_PER_SHEET = 5000


class ExcelParseError(Exception):
    """Raised when a workbook cannot be read; message is user-actionable."""


def parse_excel(data: bytes, filename: str) -> tuple[str, list[ParsedSegment], dict]:
    """Parse an .xlsx/.xls workbook into traceable segments.

    Returns (full_text, segments, metadata). Every non-empty row becomes one
    segment located at ``sheet`` / ``row N`` with a ``columns`` metadata map
    keyed by the header row (or column letters) so cell references survive
    (FR-IN-008).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ExcelParseError(
            "openpyxl is not installed; cannot parse Excel workbooks."
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParseError(
            f"'{filename}' could not be read as an Excel workbook ({exc}). "
            "If it is a legacy .xls file, re-save it as .xlsx and try again."
        ) from exc

    segments: list[ParsedSegment] = []
    text_parts: list[str] = []
    sheet_names: list[str] = []

    for ws in wb.worksheets:
        sheet_names.append(ws.title)
        rows = ws.iter_rows(values_only=True)
        header: list[str] = []
        for row_idx, row in enumerate(rows, start=1):
            if row_idx > MAX_ROWS_PER_SHEET:
                break
            cells = ["" if v is None else str(v).strip() for v in row]
            if not any(cells):
                continue
            if not header and any(cells):
                header = [c or f"col{i + 1}" for i, c in enumerate(cells)]
                continue
            columns = {
                header[i] if i < len(header) else f"col{i + 1}": cell
                for i, cell in enumerate(cells)
                if cell
            }
            line = " | ".join(f"{k}: {v}" for k, v in columns.items())
            if not line:
                continue
            text_parts.append(f"[{ws.title}] {line}")
            segments.append(
                ParsedSegment(
                    page_or_sheet=ws.title,
                    row_or_section=f"row {row_idx}",
                    content=line,
                    metadata={"columns": columns, "sheet": ws.title, "row": row_idx},
                )
            )
    wb.close()

    if not segments:
        raise ExcelParseError(
            f"'{filename}' contains no readable rows. Please provide a workbook "
            "with tabular requirement content."
        )

    metadata = {"sheets": sheet_names, "segment_count": len(segments)}
    return "\n".join(text_parts), segments, metadata
